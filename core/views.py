import uuid
import calendar
from datetime import date, datetime, timedelta

from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from django.http import HttpResponse

from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Pharmacy, Patient, Drug, Prescription, Inventory,
    PrescriptionItem,
    InventoryImportBatch, InventoryImportItem
)
from .serializers import (
    PharmacySerializer,
    PatientSerializer,
    DrugSerializer,
    PrescriptionCreateSerializer,
    PrescriptionReadSerializer,
    InventorySerializer,
)
from .permissions import IsDoctor, IsPharmacy


# =============================================================================
# Helpers
# =============================================================================

def _parse_month(month_str: str):
    """ "2026-02" -> (start_date, end_date) """
    if not month_str or len(month_str) != 7 or month_str[4] != "-":
        return None, None
    y = int(month_str[:4])
    m = int(month_str[5:7])
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, last_day)


def _current_month_str():
    now = timezone.now()
    return f"{now.year:04d}-{now.month:02d}"


def _auto_width(ws, min_w=14):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(42, max_len + 2))


def _style_header(ws, header_row=1):
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F2937")
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = align


def _latest_batch_with_items(pharmacy_id):
    """
    latest batch that has at least ONE InventoryImportItem.
    If last batch exists but empty -> ignore it.
    """
    return (
        InventoryImportBatch.objects
        .filter(pharmacy_id=pharmacy_id, items__isnull=False)
        .distinct()
        .order_by("-created_at")
        .first()
    )


def _parse_date_param(d: str, end=False):
    """
    YYYY-MM-DD -> aware datetime
    end=False -> 00:00:00
    end=True  -> 23:59:59
    """
    y, m, dd = map(int, d.split("-"))
    if end:
        return timezone.make_aware(datetime(y, m, dd, 23, 59, 59))
    return timezone.make_aware(datetime(y, m, dd, 0, 0, 0))


# =============================================================================
# Doctor utilities
# =============================================================================

@api_view(["GET"])
@permission_classes([IsDoctor])
def check_availability(request):
    pharmacy_id = request.query_params.get("pharmacy_id")
    drug_id = request.query_params.get("drug_id")

    if not pharmacy_id or not drug_id:
        return Response({"detail": "pharmacy_id and drug_id are required"}, status=400)

    inv = Inventory.objects.filter(pharmacy_id=pharmacy_id, drug_id=drug_id).first()
    qty = inv.quantity_in_stock if inv else 0
    return Response({"available_quantity": qty})


class DoctorPrescriptionHistoryView(generics.ListAPIView):
    permission_classes = [IsDoctor]
    serializer_class = PrescriptionReadSerializer

    def get_queryset(self):
        return Prescription.objects.filter(doctor=self.request.user).order_by("-created_at")


# =============================================================================
# Basic lists for Doctor UI
# =============================================================================

class PharmacyListView(generics.ListAPIView):
    permission_classes = [IsDoctor]
    queryset = Pharmacy.objects.all().order_by("name")
    serializer_class = PharmacySerializer


class PatientListView(generics.ListCreateAPIView):
    permission_classes = [IsDoctor]
    serializer_class = PatientSerializer

    def get_queryset(self):
        q = self.request.query_params.get("q", "").strip()
        qs = Patient.objects.all().order_by("full_name")
        if q:
            return qs.filter(
                Q(full_name__icontains=q) |
                Q(nic_number__icontains=q) |
                Q(phone__icontains=q)
            )
        return qs


class DrugListView(generics.ListAPIView):
    permission_classes = [IsDoctor]
    serializer_class = DrugSerializer

    def get_queryset(self):
        q = self.request.query_params.get("q", "").strip()
        qs = Drug.objects.all().order_by("generic_name")
        if q:
            return qs.filter(Q(generic_name__icontains=q) | Q(brand_name__icontains=q))
        return qs


# =============================================================================
# Prescriptions
# =============================================================================

class PrescriptionCreateView(generics.CreateAPIView):
    permission_classes = [IsDoctor]
    serializer_class = PrescriptionCreateSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        prescription = serializer.save()
        return Response(PrescriptionReadSerializer(prescription).data, status=status.HTTP_201_CREATED)


class PrescriptionQueueView(generics.ListAPIView):
    permission_classes = [IsPharmacy]
    serializer_class = PrescriptionReadSerializer

    def get_queryset(self):
        status_filter = (self.request.query_params.get("status") or "").strip().upper()

        profile = getattr(self.request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Prescription.objects.none()

        qs = Prescription.objects.filter(pharmacy_id=profile.pharmacy_id).order_by("-created_at")
        return qs.filter(status=status_filter) if status_filter else qs.filter(status="PENDING")


class PrescriptionStatusUpdateView(generics.GenericAPIView):
    permission_classes = [IsPharmacy]

    def patch(self, request, pk):
        status_to = (request.data.get("status") or "").strip().upper()
        reason = (request.data.get("reason") or "").strip()

        if status_to not in ["READY", "CANCELLED"]:
            return Response({"detail": "Invalid status. Use READY or CANCELLED."}, status=400)

        try:
            prescription = Prescription.objects.select_related("pharmacy").get(prescription_id=pk)
        except Prescription.DoesNotExist:
            return Response({"detail": "Prescription not found"}, status=404)

        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        if str(profile.pharmacy_id) != str(prescription.pharmacy_id):
            return Response({"detail": "Forbidden for this pharmacy"}, status=403)

        if status_to == "READY":
            if prescription.status != "PENDING":
                return Response({"detail": f"Cannot mark READY from {prescription.status}"}, status=400)
            prescription.status = "READY"

        if status_to == "CANCELLED":
            if prescription.status == "DISPENSED":
                return Response({"detail": "Cannot cancel a DISPENSED prescription"}, status=400)
            if prescription.status == "CANCELLED":
                return Response({"detail": "Already CANCELLED"}, status=400)
            prescription.status = "CANCELLED"
            if hasattr(prescription, "cancel_reason"):
                prescription.cancel_reason = reason

        prescription.save()
        return Response(PrescriptionReadSerializer(prescription).data, status=200)


class DispensePrescriptionView(generics.GenericAPIView):
    permission_classes = [IsPharmacy]

    def post(self, request, pk):
        try:
            prescription = Prescription.objects.select_related("pharmacy").get(prescription_id=pk)
        except Prescription.DoesNotExist:
            return Response({"detail": "Prescription not found"}, status=404)

        profile = getattr(request.user, "profile", None)
        if not profile or not profile.pharmacy_id or str(profile.pharmacy_id) != str(prescription.pharmacy_id):
            return Response({"detail": "Forbidden for this pharmacy"}, status=403)

        if prescription.status in ["DISPENSED", "CANCELLED"]:
            return Response({"detail": f"Cannot dispense. Current status: {prescription.status}"}, status=400)

        with transaction.atomic():
            items = prescription.items.select_related("drug").all()

            for it in items:
                inv, _ = Inventory.objects.select_for_update().get_or_create(
                    pharmacy=prescription.pharmacy,
                    drug=it.drug,
                    defaults={"quantity_in_stock": 0, "reorder_level": 0},
                )

                if inv.quantity_in_stock < it.quantity:
                    return Response(
                        {"detail": f"Insufficient stock for {it.drug}. Have {inv.quantity_in_stock}, need {it.quantity}"},
                        status=400,
                    )

                inv.quantity_in_stock -= it.quantity
                inv.save()

            prescription.status = "DISPENSED"
            prescription.dispensed_at = timezone.now()
            prescription.dispensed_by = request.user
            prescription.save()

        return Response(PrescriptionReadSerializer(prescription).data, status=200)


class PrescriptionDetailView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PrescriptionReadSerializer
    lookup_field = "prescription_id"
    lookup_url_kwarg = "pk"

    def get_queryset(self):
        user = self.request.user
        profile = getattr(user, "profile", None)

        qs = Prescription.objects.all()
        if not profile:
            return Prescription.objects.none()

        if profile.role == "DOCTOR":
            return qs.filter(doctor=user)

        if profile.role == "PHARMACY":
            if not profile.pharmacy_id:
                return Prescription.objects.none()
            return qs.filter(pharmacy_id=profile.pharmacy_id)

        return qs


# =============================================================================
# Pharmacy Inventory
# =============================================================================

class PharmacyInventoryListView(generics.ListAPIView):
    permission_classes = [IsPharmacy]
    serializer_class = InventorySerializer

    def get_queryset(self):
        profile = getattr(self.request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Inventory.objects.none()

        return (
            Inventory.objects
            .filter(pharmacy_id=profile.pharmacy_id)
            .select_related("drug")
            .order_by("drug__generic_name")
        )


class PharmacyInventoryUpdateView(generics.UpdateAPIView):
    permission_classes = [IsPharmacy]
    serializer_class = InventorySerializer
    lookup_field = "inventory_id"
    queryset = Inventory.objects.select_related("pharmacy", "drug").all()

    def partial_update(self, request, *args, **kwargs):
        inv = self.get_object()

        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        if str(inv.pharmacy_id) != str(profile.pharmacy_id):
            return Response({"detail": "Forbidden for this pharmacy"}, status=403)

        allowed = {"quantity_in_stock", "reorder_level", "expiry_date"}
        clean_data = {k: v for k, v in request.data.items() if k in allowed}

        serializer = self.get_serializer(inv, data=clean_data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data, status=200)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


# =============================================================================
# ✅ EXPORT — Last Imported vs Current vs Reduced (with fallback)
# =============================================================================

class InventoryExportView(APIView):
    """
    GET /api/inventory/export/

    ✅ DEFAULT:
      Brand | Generic | Category | Last Imported Qty | Current Qty | Reduced Qty

    Optional:
      ?mode=basic   -> old 5 columns
    """
    permission_classes = [IsPharmacy]

    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        mode = (request.query_params.get("mode") or "").strip().lower()

        # -------------------------------
        # BASIC export
        # -------------------------------
        if mode == "basic":
            qs = (
                Inventory.objects
                .filter(pharmacy_id=profile.pharmacy_id)
                .select_related("drug")
                .order_by("drug__generic_name")
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory List"

            headers = ["Brand Name", "Generic Name", "Category", "Quantity in Stock", "Reorder Level"]
            ws.append(headers)

            for inv in qs:
                ws.append([
                    inv.drug.brand_name or "",
                    inv.drug.generic_name or "",
                    inv.drug.category or "",
                    int(inv.quantity_in_stock or 0),
                    int(inv.reorder_level or 0),
                ])

            _style_header(ws, 1)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            _auto_width(ws, 18)

            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="pharmacy_inventory_basic.xlsx"'
            wb.save(resp)
            return resp

        # -------------------------------
        # DEFAULT export you asked
        # -------------------------------
        inv_qs = (
            Inventory.objects
            .filter(pharmacy_id=profile.pharmacy_id)
            .select_related("drug")
            .order_by("drug__generic_name")
        )
        current_map = {str(inv.drug_id): int(inv.quantity_in_stock or 0) for inv in inv_qs}

        last_batch = _latest_batch_with_items(profile.pharmacy_id)

        last_import_map = {}
        last_import_date = None

        if last_batch:
            last_import_date = last_batch.created_at
            for drug_id, qty in (
                InventoryImportItem.objects
                .filter(batch=last_batch)
                .values_list("drug_id", "quantity_in_stock")
            ):
                last_import_map[str(drug_id)] = int(qty or 0)

        # ✅ KEY FIX: if snapshot missing => avoid 0 by using current as last-import baseline
        if not last_import_map:
            last_import_map = dict(current_map)

        all_drug_ids = set(last_import_map.keys()) | set(current_map.keys())

        drugs = Drug.objects.filter(drug_id__in=list(all_drug_ids))
        drug_map = {str(d.drug_id): d for d in drugs}

        def sort_key(did):
            d = drug_map.get(did)
            if not d:
                return ("", "")
            return ((d.generic_name or "").lower(), (d.brand_name or "").lower())

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Summary"

        headers = ["Brand Name", "Generic Name", "Category", "Last Imported Qty", "Current Qty", "Reduced Qty"]
        ws.append(headers)

        for did in sorted(all_drug_ids, key=sort_key):
            d = drug_map.get(did)
            if not d:
                continue

            last_qty = int(last_import_map.get(did, 0))
            cur_qty = int(current_map.get(did, 0))

            reduced = last_qty - cur_qty
            if reduced < 0:
                reduced = 0

            ws.append([
                d.brand_name or "",
                d.generic_name or "",
                d.category or "",
                last_qty,
                cur_qty,
                reduced
            ])

        _style_header(ws, 1)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _auto_width(ws, 18)

        filename = "inventory_stock_summary.xlsx"
        if last_import_date:
            filename = f"inventory_stock_summary_{last_import_date.strftime('%Y-%m-%d')}.xlsx"

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


# =============================================================================
# Import view
# =============================================================================

class InventoryImportView(APIView):
    """
    POST /api/inventory/import/
    Upload Excel and:
    - upsert drugs
    - update inventory
    - store batch snapshot items (important for export last imported qty)
    """
    permission_classes = [IsPharmacy]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Excel file is required (field name: file)"}, status=400)

        wb = load_workbook(filename=f, data_only=True)
        ws = wb.active

        raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        norm = [" ".join(h.replace("\n", " ").replace("\r", " ").split()).lower() for h in raw_headers]

        def idx(name):
            n = name.lower()
            return norm.index(n) if n in norm else -1

        i_brand = idx("brand name")
        i_generic = idx("generic name")
        i_category = idx("category")
        i_qty = idx("quantity in stock")
        i_reorder = idx("reorder level")

        if i_generic == -1 or i_qty == -1:
            return Response(
                {"detail": "Invalid template. Required columns: Generic Name, Quantity in Stock"},
                status=400
            )

        updated_inventory = 0
        created_drugs = 0
        skipped = 0
        errors = []

        with transaction.atomic():
            batch = InventoryImportBatch.objects.create(
                pharmacy_id=profile.pharmacy_id,
                imported_by=request.user,
                file_name=getattr(f, "name", ""),
            )

            for r in range(2, ws.max_row + 1):
                brand = ws.cell(row=r, column=i_brand + 1).value if i_brand != -1 else ""
                generic = ws.cell(row=r, column=i_generic + 1).value
                category = ws.cell(row=r, column=i_category + 1).value if i_category != -1 else ""
                qty = ws.cell(row=r, column=i_qty + 1).value
                reorder = ws.cell(row=r, column=i_reorder + 1).value if i_reorder != -1 else 0

                brand = (str(brand).strip() if brand is not None else "")
                generic = (str(generic).strip() if generic is not None else "")
                category = (str(category).strip() if category is not None else "")

                if not generic:
                    skipped += 1
                    continue

                try:
                    qty_int = int(float(qty)) if qty is not None and str(qty).strip() != "" else 0
                    reorder_int = int(float(reorder)) if reorder is not None and str(reorder).strip() != "" else 0
                except Exception:
                    skipped += 1
                    errors.append(f"Row {r}: qty/reorder not numeric")
                    continue

                if brand:
                    drug = Drug.objects.filter(generic_name__iexact=generic, brand_name__iexact=brand).first()
                else:
                    drug = Drug.objects.filter(generic_name__iexact=generic).first()

                if not drug:
                    drug = Drug.objects.create(
                        generic_name=generic,
                        brand_name=brand,
                        category=category,
                        unit="",
                        requires_prescription=True,
                        description="",
                    )
                    created_drugs += 1
                else:
                    if category and not (drug.category or "").strip():
                        drug.category = category
                        drug.save()

                inv, _ = Inventory.objects.select_for_update().get_or_create(
                    pharmacy_id=profile.pharmacy_id,
                    drug_id=drug.drug_id,
                    defaults={"quantity_in_stock": 0, "reorder_level": 0},
                )
                inv.quantity_in_stock = qty_int
                inv.reorder_level = reorder_int
                inv.save()
                updated_inventory += 1

                InventoryImportItem.objects.update_or_create(
                    batch=batch,
                    drug=drug,
                    defaults={"quantity_in_stock": qty_int, "reorder_level": reorder_int}
                )

        return Response({
            "detail": "Import complete",
            "batch_id": str(batch.batch_id),
            "updated_inventory": updated_inventory,
            "created_drugs": created_drugs,
            "skipped": skipped,
            "errors": errors[:30],
        }, status=200)


# =============================================================================
# ✅ NEW: Dispensed items Excel report
# =============================================================================

class DispensedReportExportView(APIView):
    """
    GET /api/reports/dispensed/export/?from=YYYY-MM-DD&to=YYYY-MM-DD
    Pharmacy only.

    Excel columns:
    Dispensed DateTime | Prescription ID | Drug (Generic) | Brand | Category | Qty | Patient | Dispensed By
    """
    permission_classes = [IsPharmacy]

    def get(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        df = (request.query_params.get("from") or "").strip()
        dt = (request.query_params.get("to") or "").strip()

        # default last 30 days
        end_dt = timezone.now()
        start_dt = end_dt - timedelta(days=30)

        try:
            if df:
                start_dt = _parse_date_param(df, end=False)
            if dt:
                end_dt = _parse_date_param(dt, end=True)
        except Exception:
            return Response({"detail": "Invalid date format. Use from=YYYY-MM-DD&to=YYYY-MM-DD"}, status=400)

        qs = (
            PrescriptionItem.objects
            .select_related(
                "drug",
                "prescription",
                "prescription__patient",
                "prescription__dispensed_by"
            )
            .filter(
                prescription__pharmacy_id=profile.pharmacy_id,
                prescription__status="DISPENSED",
                prescription__dispensed_at__gte=start_dt,
                prescription__dispensed_at__lte=end_dt,
            )
            .order_by("-prescription__dispensed_at")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Dispensed Log"

        headers = [
            "Dispensed DateTime",
            "Prescription ID",
            "Generic Name",
            "Brand Name",
            "Category",
            "Quantity",
            "Patient",
            "Dispensed By"
        ]
        ws.append(headers)

        for it in qs:
            p = it.prescription
            drug = it.drug
            patient_name = p.patient.full_name if p.patient else ""
            pharmacist = ""
            if p.dispensed_by:
                pharmacist = p.dispensed_by.get_full_name() or p.dispensed_by.username

            ws.append([
                p.dispensed_at.strftime("%Y-%m-%d %H:%M:%S") if p.dispensed_at else "",
                str(p.prescription_id),
                drug.generic_name or "",
                drug.brand_name or "",
                drug.category or "",
                int(it.quantity or 0),
                patient_name,
                pharmacist,
            ])

        _style_header(ws, 1)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _auto_width(ws, 18)

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="dispensed_items_log.xlsx"'
        wb.save(resp)
        return resp


# =============================================================================
# Compatibility endpoint (if you still have this URL in urls.py)
# =============================================================================

class InventoryMonthlyReportExportView(APIView):
    permission_classes = [IsPharmacy]

    def get(self, request):
        return InventoryExportView().get(request)




import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Inventory, InventoryImportItem, Prescription
from .utils import model  


def normalize(s):
    """Lowercase + strip for fuzzy brand name matching."""
    return str(s or "").lower().strip()


class PharmacyPredictionView(APIView):
    permission_classes = [IsAuthenticated]

    def get_profile(self, request):
        profile = getattr(request.user, "profile", None)
        if not profile or profile.role != "PHARMACY" or not profile.pharmacy_id:
            return None
        return profile

    def get_db_sales(self, profile, now):
        """
        Build monthly_sales[drug_id_str][m_offset] = qty
        from DISPENSED prescriptions in the DB for last 12 months.
        """
        monthly_sales = {}

        for m_offset in range(1, 13):
            ref = now - relativedelta(months=m_offset)
            prescriptions = Prescription.objects.filter(
                pharmacy_id=profile.pharmacy_id,
                status="DISPENSED",
                dispensed_at__year=ref.year,
                dispensed_at__month=ref.month,
            ).prefetch_related("items__drug")

            for rx in prescriptions:
                for item in rx.items.all():
                    did = str(item.drug.drug_id)
                    if did not in monthly_sales:
                        monthly_sales[did] = {}
                    monthly_sales[did][m_offset] = monthly_sales[did].get(m_offset, 0) + item.quantity

        return monthly_sales

    def parse_uploaded_sales(self, file):
        """
        Parse uploaded CSV/Excel and return { normalized_brand_name: qty }.
        Raises ValueError with a user-friendly message for bad/empty files.
        """
       
        try:
            if file.name.lower().endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            raise ValueError(f"Could not read file. Make sure it is a valid CSV or Excel file. ({e})")

        
        if df.empty or len(df.columns) == 0:
            raise ValueError("The uploaded file is empty. Please upload a file with sales data.")

        if len(df) == 0:
            raise ValueError("The uploaded file has no data rows. Please upload a file with at least one sale record.")

        
        name_col = None
        for col in df.columns:
            if col.strip().lower() in ["brand name", "brand_name", "drug", "drug name", "drug_name", "name"]:
                name_col = col
                break

        if name_col is None:
            raise ValueError(
                f"Wrong file format — could not find a brand name column. "
                f"Expected one of: 'Brand Name', 'Drug Name', 'Name'. "
                f"Found columns: {', '.join(str(c) for c in df.columns)}"
            )

        
        qty_col = None
        for col in df.columns:
            if col.strip().lower() in ["quantity", "qty", "amount", "sold", "units"]:
                qty_col = col
                break

        if qty_col is None:
            raise ValueError(
                f"Wrong file format — could not find a quantity column. "
                f"Expected one of: 'Quantity', 'Qty', 'Amount', 'Sold', 'Units'. "
                f"Found columns: {', '.join(str(c) for c in df.columns)}"
            )

        
        valid_rows = df[name_col].notna() & df[qty_col].notna()
        if valid_rows.sum() == 0:
            raise ValueError(
                "Wrong file format — the file has columns but all rows are empty or blank. "
                "Please upload a file with actual sales data."
            )

        
        sales = {}
        for _, row in df[valid_rows].iterrows():
            name = normalize(row[name_col])
            try:
                qty = int(float(row[qty_col] or 0))
            except (ValueError, TypeError):
                continue  
            if name:
                sales[name] = sales.get(name, 0) + qty

        if not sales:
            raise ValueError(
                "Wrong file format — no valid sales records could be read. "
                "Make sure Brand Name and Quantity columns have real values."
            )

        return sales

    def build_features(self, profile, now, uploaded_sales=None):
        """
        Build the 30 XGBoost features, one row per drug in inventory.
        LAG_1 priority:
          1. Uploaded file (matched by normalized brand name)
          2. DB dispensed prescriptions last month
        """
        next_month_dt = now + relativedelta(months=1)
        month   = next_month_dt.month
        year    = next_month_dt.year
        quarter = (month - 1) // 3 + 1

        is_monsoon    = 1 if month in [5, 6, 7, 8, 9, 10] else 0
        is_year_end   = 1 if month == 12 else 0
        is_year_start = 1 if month == 1  else 0
        is_avurudu    = 1 if month == 4  else 0
        is_vesak      = 1 if month == 5  else 0
        is_deepavali  = 1 if month == 10 else 0
        month_sin = np.sin(2 * np.pi * month / 12)
        month_cos = np.cos(2 * np.pi * month / 12)

        db_sales = self.get_db_sales(profile, now)

        inventory_qs = Inventory.objects.filter(
            pharmacy_id=profile.pharmacy_id
        ).select_related("drug")

        category_map = {
            "Tablet": 1, "Capsule": 2, "Injection": 3, "Syrup": 4,
            "Cream": 5, "Drops": 6, "Inhaler": 7, "Patch": 8,
        }

        rows = []
        meta = []

        for inv in inventory_qs:
            drug          = inv.drug
            drug_id_str   = str(drug.drug_id)
            drug_id_int   = drug.drug_id.int % 10**9
            category_id   = category_map.get(drug.category, 0)
            current_stock = inv.quantity_in_stock
            brand_norm    = normalize(drug.brand_name)

            drug_sales = db_sales.get(drug_id_str, {})
            lag1  = drug_sales.get(1,  0)
            lag2  = drug_sales.get(2,  0)
            lag3  = drug_sales.get(3,  0)
            lag6  = drug_sales.get(6,  0)
            lag12 = drug_sales.get(12, 0)

        
            if uploaded_sales is not None:
                if brand_norm in uploaded_sales:
                    lag1 = uploaded_sales[brand_norm]
                else:
                    for uname, uqty in uploaded_sales.items():
                        if uname in brand_norm or brand_norm in uname:
                            lag1 = uqty
                            break

            recent3  = [drug_sales.get(i, 0) for i in range(1, 4)];  recent3[0]  = lag1
            recent6  = [drug_sales.get(i, 0) for i in range(1, 7)];  recent6[0]  = lag1
            recent12 = [drug_sales.get(i, 0) for i in range(1, 13)]; recent12[0] = lag1

            roll3_mean  = float(np.mean(recent3))
            roll6_mean  = float(np.mean(recent6))
            roll12_mean = float(np.mean(recent12))
            roll3_std   = float(np.std(recent3))
            roll3_max   = float(np.max(recent3))
            roll3_min   = float(np.min(recent3))
            mom_change  = lag1 - lag2
            yoy_change  = lag1 - lag12

            unit_price    = float(getattr(drug, "unit_price_lkr", 0) or 0)
            reorder_level = int(getattr(inv, "reorder_level", 0) or 0)

            last_item = InventoryImportItem.objects.filter(
                batch__pharmacy_id=profile.pharmacy_id,
                drug=drug
            ).order_by("-batch__created_at").first()

            if last_item and last_item.quantity_in_stock > 0:
                last_qty = last_item.quantity_in_stock
            else:
                last_qty = current_stock + lag1

            reduced       = lag1
            stock_ratio   = (current_stock / last_qty) if last_qty > 0 else 1.0
            stock_cushion = (current_stock - reorder_level) if reorder_level > 0 else current_stock

            rows.append({
                "DRUG_ID":        drug_id_int,
                "CATEGORY_ID":    category_id,
                "MONTH":          month,
                "YEAR":           year,
                "QUARTER":        quarter,
                "MONTH_SIN":      month_sin,
                "MONTH_COS":      month_cos,
                "IS_MONSOON":     is_monsoon,
                "IS_YEAR_END":    is_year_end,
                "IS_YEAR_START":  is_year_start,
                "IS_AVURUDU":     is_avurudu,
                "IS_VESAK":       is_vesak,
                "IS_DEEPAVALI":   is_deepavali,
                "LAG_1":          lag1,
                "LAG_2":          lag2,
                "LAG_3":          lag3,
                "LAG_6":          lag6,
                "LAG_12":         lag12,
                "ROLL_3_MEAN":    roll3_mean,
                "ROLL_6_MEAN":    roll6_mean,
                "ROLL_12_MEAN":   roll12_mean,
                "ROLL_3_STD":     roll3_std,
                "ROLL_3_MAX":     roll3_max,
                "ROLL_3_MIN":     roll3_min,
                "MOM_CHANGE":     mom_change,
                "YOY_CHANGE":     yoy_change,
                "UNIT_PRICE_LKR": unit_price,
                "REORDER_LEVEL":  reorder_level,
                "STOCK_RATIO":    stock_ratio,
                "STOCK_CUSHION":  stock_cushion,
            })

            meta.append({
                "brand_name":   drug.brand_name,
                "generic_name": drug.generic_name,
                "category":     drug.category,
                "last":         last_qty,
                "current":      current_stock,
                "reduced":      reduced,
            })

        return pd.DataFrame(rows), meta

    def _run_prediction(self, profile, now, uploaded_sales=None):
        next_month_str = (now + relativedelta(months=1)).strftime("%B %Y")
        feature_df, meta = self.build_features(profile, now, uploaded_sales)

        if feature_df.empty:
            return Response({"next_month": next_month_str, "data": []})

        try:
            predicted_qty = model.predict(feature_df)
        except Exception as e:
            return Response({"detail": f"Prediction error: {str(e)}"}, status=500)

        result = [
            {**m, "predicted": max(int(predicted_qty[i]), 0)}
            for i, m in enumerate(meta)
        ]
        return Response({"next_month": next_month_str, "data": result})

    def get(self, request):
        now = timezone.now()
        profile = self.get_profile(request)
        if not profile:
            return Response({"detail": "Pharmacy profile missing"}, status=403)
        return self._run_prediction(profile, now, uploaded_sales=None)

    def post(self, request):
        now = timezone.now()
        profile = self.get_profile(request)
        if not profile:
            return Response({"detail": "Pharmacy profile missing"}, status=403)

        uploaded_sales = None
        if "file" in request.FILES:
            try:
                uploaded_sales = self.parse_uploaded_sales(request.FILES["file"])
            except ValueError as e:
                return Response({"detail": str(e)}, status=400)

        return self._run_prediction(profile, now, uploaded_sales)